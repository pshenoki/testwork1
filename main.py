from openpyxl import load_workbook
from models import Task1, Task2, Task3, Task3mod, Task4, Task5, Task6

FILENAME = 'task_support.xlsx'
SHEET = 'Tasks'
START_BORD = 3
FINISH_BORD = 1000

wb = load_workbook(FILENAME)
ws = wb[SHEET]

for row in range(START_BORD, FINISH_BORD):
    task1 = ws[row][1].value
    Task1.check(task1)
    task2 = ws[row][2].value
    Task2.check(task2)
    task3 = ws[row][3].value
    Task3.check(task3)
    Task3mod.check(task3)
    task4 = ws[row][4].value
    Task4.check(task4)
    task5 = ws[row][5].value
    Task5.check(task5)
    task6 = ws[row][6].value
    Task6.check(task6)

Task1.info()
Task2.info()
Task3.info()
Task3mod.info()
Task4.info()
Task5.info()
Task6.info()
